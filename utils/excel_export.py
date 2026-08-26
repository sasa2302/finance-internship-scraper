"""Generation et relecture du classeur Excel unique.

Un seul fichier : data/stages_finance_marche.xlsx, reecrit a chaque run.

  - Resume     : compteurs du jour et de la fenetre courante
  - Off-Cycle  : stages longs / cesure
  - Summer     : programmes d'ete
  - A trier    : calendrier non identifie
  - _donnees   : feuille technique masquee, relue au run suivant pour
                 conserver les offres des jours precedents

Les offres du jour sont en tete et marquees "Nouveau".
"""

import logging
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Schema technique (feuille masquee) : sert de base de donnees entre les runs
DATA_FIELDS = [
    "id", "title", "company", "employer_category", "location", "zone",
    "zone_label", "period_label", "period_note", "internship_type",
    "type_reason", "duration", "url", "date_posted", "date_added",
    "source", "relevance_score", "description_snippet",
]

# (en-tete affiche, largeur, cle)
COLUMNS = [
    ("Nouveau", 10, "_is_new"),
    ("Poste", 50, "title"),
    ("Entreprise", 24, "company"),
    ("Type employeur", 22, "employer_category"),
    ("Lieu", 26, "location"),
    ("Zone", 15, "zone_label"),
    ("Duree", 13, "duration"),
    ("Periode visee", 15, "_period"),
    ("Pourquoi ce type", 28, "type_reason"),
    ("Score", 8, "relevance_score"),
    ("Publiee le", 13, "date_posted"),
    ("Ajoutee le", 13, "date_added"),
    ("Source", 13, "source"),
    ("Lien", 14, "url"),
    ("Description", 65, "description_snippet"),
]

DATA_SHEET = "_donnees"
SHEETS = [
    ("Off-Cycle", "off_cycle", "OFF-CYCLE - stages longs (4-6 mois), cesure, stage de fin d'etudes"),
    ("Summer", "summer", "SUMMER - programmes d'ete (Summer Analyst, 8-12 semaines)"),
    ("A trier", "unknown", "A TRIER - calendrier non identifie automatiquement"),
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
LINK_FONT = Font(color="0563C1", underline="single")
NEW_FILL = PatternFill("solid", fgColor="FFF2CC")
NEW_FONT = Font(bold=True, color="BF8F00")

SCORE_FILLS = [
    (0.75, PatternFill("solid", fgColor="C6EFCE")),
    (0.50, PatternFill("solid", fgColor="FFEB9C")),
    (0.00, PatternFill("solid", fgColor="FFC7CE")),
]

ZONE_ORDER = {"CORE": 0, "EUROPE": 1, "GLOBAL": 2, "INCONNU": 3}
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _score_fill(score):
    for threshold, fill in SCORE_FILLS:
        if score >= threshold:
            return fill
    return SCORE_FILLS[-1][1]


def read_existing(path):
    """Relit la feuille technique du classeur precedent.

    Renvoie une liste de dictionnaires, vide si le fichier n'existe pas ou
    n'est pas exploitable (on repart alors de zero plutot que d'echouer).
    """
    path = Path(path)
    if not path.exists():
        return []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        if DATA_SHEET not in wb.sheetnames:
            return []
        ws = wb[DATA_SHEET]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h) if h is not None else "" for h in rows[0]]
        out = []
        for raw in rows[1:]:
            row = {h: ("" if v is None else v) for h, v in zip(headers, raw)}
            if row.get("url"):
                out.append(row)
        wb.close()
        return out
    except Exception as e:
        logger.warning(f"Classeur precedent illisible ({e}), on repart de zero.")
        return []


def _cell_value(row, key):
    if key == "_is_new":
        return "NOUVEAU" if row.get("_is_new") else ""
    if key == "_period":
        label = str(row.get("period_label") or "")
        note = str(row.get("period_note") or "")
        return label or ("a confirmer" if note else "")
    value = row.get(key, "")
    if key == "relevance_score":
        try:
            return round(float(value), 2)
        except (TypeError, ValueError):
            return 0.0
    return "" if value is None else str(value)


def _write_sheet(ws, rows, sheet_title):
    ws.freeze_panes = "A3"
    new_count = sum(1 for r in rows if r.get("_is_new"))

    ws.cell(row=1, column=1, value=sheet_title).font = TITLE_FONT
    ws.cell(row=1, column=len(COLUMNS),
            value=f"{len(rows)} offre(s) - dont {new_count} nouvelle(s)").alignment = (
        Alignment(horizontal="right"))

    for idx, (header, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=idx, value=header)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[2].height = 28

    # Nouveautes d'abord, puis zone prioritaire, puis score
    ordered = sorted(rows, key=lambda r: (
        0 if r.get("_is_new") else 1,
        ZONE_ORDER.get(str(r.get("zone") or "INCONNU"), 9),
        -float(r.get("relevance_score") or 0),
        str(r.get("company") or ""),
    ))

    for i, row in enumerate(ordered, start=3):
        for c, (_, _, key) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=i, column=c, value=_cell_value(row, key))
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=(key in ("title", "description_snippet")))
            if key == "url" and row.get("url"):
                cell.value = "Voir l'offre"
                cell.hyperlink = str(row["url"])
                cell.font = LINK_FONT
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif key == "relevance_score":
                cell.fill = _score_fill(float(row.get("relevance_score") or 0))
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif key == "_is_new" and row.get("_is_new"):
                cell.fill, cell.font = NEW_FILL, NEW_FONT
                cell.alignment = Alignment(horizontal="center", vertical="top")

    if ordered:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{len(ordered) + 2}"
    else:
        ws.cell(row=3, column=1, value="Aucune offre de ce type dans la fenetre courante.")


def _write_summary(ws, buckets, run_stats, window_days):
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16
    row = 1
    ws.cell(row=row, column=1, value="Stages Finance de Marche").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1,
            value=f"Mis a jour le {datetime.now(timezone.utc).strftime('%d/%m/%Y a %H:%M UTC')}")
    row += 1
    ws.cell(row=row, column=1, value=f"Offres conservees {window_days} jours - "
                                     f"les nouveautes du jour sont en tete de chaque onglet")
    row += 2

    def section(title, pairs):
        nonlocal row
        c = ws.cell(row=row, column=1, value=title)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        ws.cell(row=row, column=2).fill = HEADER_FILL
        row += 1
        for label, value in pairs:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value).alignment = Alignment(horizontal="right")
            row += 1
        row += 1

    allr = [r for v in buckets.values() for r in v]
    new = [r for r in allr if r.get("_is_new")]

    section("Nouveautes de ce run", [
        ("Off-Cycle", sum(1 for r in buckets["off_cycle"] if r.get("_is_new"))),
        ("Summer", sum(1 for r in buckets["summer"] if r.get("_is_new"))),
        ("A trier", sum(1 for r in buckets["unknown"] if r.get("_is_new"))),
        ("TOTAL nouveau", len(new)),
    ])
    section(f"Total sur {window_days} jours", [
        ("Off-Cycle", len(buckets["off_cycle"])),
        ("Summer", len(buckets["summer"])),
        ("A trier", len(buckets["unknown"])),
        ("TOTAL", len(allr)),
    ])

    zones = {}
    for r in allr:
        k = str(r.get("zone_label") or "Non precise")
        zones[k] = zones.get(k, 0) + 1
    section("Repartition geographique", sorted(zones.items(), key=lambda kv: -kv[1])[:12])

    cats = {}
    for r in allr:
        k = str(r.get("employer_category") or "Non classe")
        cats[k] = cats.get(k, 0) + 1
    section("Type d'employeur", sorted(cats.items(), key=lambda kv: -kv[1]))

    if run_stats:
        section("Statistiques du run", list(run_stats.items()))


def _write_data_sheet(ws, buckets):
    ws.append(DATA_FIELDS)
    for bucket in buckets.values():
        for row in bucket:
            ws.append([row.get(f, "") for f in DATA_FIELDS])
    ws.sheet_state = "hidden"


def build_workbook(buckets, output_path, run_stats=None, window_days=60):
    """Ecrit le classeur unique. `buckets` : dict type -> liste de dictionnaires."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Resume"
    _write_summary(ws, buckets, run_stats, window_days)

    for sheet_name, key, title in SHEETS:
        _write_sheet(wb.create_sheet(sheet_name), buckets[key], title)

    _write_data_sheet(wb.create_sheet(DATA_SHEET), buckets)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info(f"Classeur ecrit : {output_path}")
    return output_path
